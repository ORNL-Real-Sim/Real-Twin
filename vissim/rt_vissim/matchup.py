##############################################################################
# Copyright (c) 2024-, Oak Ridge National Laboratory                          #
# All rights reserved.                                                       #
#                                                                            #
# This file is part of RealTwin and is distributed under a GPL               #
# license. For the licensing terms see the LICENSE file in the top-level     #
# directory.                                                                 #
#                                                                            #
# Contributors: ORNL Real-Twin Team                                          #
# Contact: realtwin@ornl.gov                                                 #
##############################################################################
"""Generate and read the Vissim MatchupTable.

The MatchupTable is RealTwin's single hand-curated artefact: one row per
(junction, approach, turn), tying three namespaces together.

===========================  ==============================================
Column group                 What it identifies
===========================  ==============================================
``*_Vissim``                 The network: junction, from/to link numbers
``*_GridSmart``              The demand: turn-count file, intersection, turn
``*_Synchro``                The control: UTDF file, INTID, turn
===========================  ==============================================

The layout mirrors
:func:`~realtwin.func_lib._c_abstract_scenario.rt_matchup_table_generation.generate_matchup_table`
-- a merged ``Network`` / ``Demand`` / ``Signal`` banner row, the real header
underneath, and merged cells for the values that are constant within a junction.
The network columns differ: Vissim link numbers replace OpenDRIVE road IDs, and
an extra column records the junction-internal links a movement traverses, which
signal-head placement needs later.

As in the SUMO flow, the demand and signal columns are written **blank** for the
user to fill in.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment

#: Network columns, sourced from :func:`rt_vissim.network.build_movement_table`.
#: Positionally identical to the SUMO table's network block, with link numbers
#: substituted for road IDs.
NETWORK_COLUMNS = ["JunctionID_Vissim", "Bearing", "Numbering", "FromLinkNo_Vissim",
                   "ToLinkNo_Vissim", "Turn"]

#: Demand columns, filled in by the user.
DEMAND_COLUMNS = ["File_GridSmart", "Date_GridSmart", "IntersectionName_GridSmart",
                  "Turn_GridSmart"]

#: Signal columns, filled in by the user.
SIGNAL_COLUMNS = ["File_Synchro", "IntersectionID_Synchro", "Turn_Synchro"]

#: Trailing columns that belong to no group.
OTHER_COLUMNS = ["Need calibration?"]

#: Vissim-only columns, parked past the end of the SUMO layout so that columns
#: A..N stay positionally identical to the SUMO MatchupTable.  Signal-head
#: placement needs the junction-internal links a movement traverses.
EXTRA_COLUMNS = ["InternalLinks_Vissim"]

#: Columns 1..14 mirror the SUMO table exactly; column 15 is the Vissim extra.
ALL_COLUMNS = (NETWORK_COLUMNS + DEMAND_COLUMNS + SIGNAL_COLUMNS
               + OTHER_COLUMNS + EXTRA_COLUMNS)

#: The movement-table columns this module consumes.
REQUIRED_COLUMNS = NETWORK_COLUMNS + EXTRA_COLUMNS

#: Columns merged vertically across the rows of one junction, and therefore
#: forward-filled on read.  ``File_Synchro`` is absent by design: as in the SUMO
#: table it is merged across the whole sheet, since one UTDF export covers the
#: entire network.
MERGED_COLUMNS = ["JunctionID_Vissim", "File_GridSmart", "Date_GridSmart",
                  "IntersectionName_GridSmart", "IntersectionID_Synchro",
                  "Need calibration?"]

#: Columns forward-filled on read: the per-junction merges plus the sheet-wide
#: ``File_Synchro``.
FILLED_COLUMNS = MERGED_COLUMNS + ["File_Synchro"]

#: Column widths, keyed by letter, matching the SUMO table's readability.
COLUMN_WIDTHS = {"A": 20, "B": 12, "C": 12, "D": 20, "E": 20, "F": 12, "G": 22,
                 "H": 16, "I": 28, "J": 16, "K": 22, "L": 22, "M": 14, "N": 18,
                 "O": 32}


def generate_matchup_table(movements: pd.DataFrame,
                           path_output: str | Path = "MatchupTable.xlsx") -> Path:
    """Write a blank-to-fill Vissim MatchupTable from a movement table.

    Args:
        movements: Output of
            :func:`rt_vissim.network.build_movement_table`.  Must carry the
            columns in :data:`NETWORK_COLUMNS`; anything else is ignored.
        path_output: Destination ``.xlsx`` path.

    Returns:
        The written path.

    Raises:
        ValueError: If ``movements`` is empty or missing network columns.
    """
    if movements is None or movements.empty:
        raise ValueError("No movements to write; check the network extraction stage.")
    missing = [c for c in REQUIRED_COLUMNS if c not in movements.columns]
    if missing:
        raise ValueError(f"Movement table is missing columns: {missing}")

    path_output = Path(path_output)
    path_output.parent.mkdir(parents=True, exist_ok=True)

    df = movements[REQUIRED_COLUMNS].copy()
    for col in DEMAND_COLUMNS + SIGNAL_COLUMNS + OTHER_COLUMNS:
        df[col] = None
    df = df[ALL_COLUMNS]
    return write_matchup_table(df, path_output)


def write_matchup_table(df: pd.DataFrame,
                        path_output: str | Path = "MatchupTable.xlsx") -> Path:
    """Write a full MatchupTable frame to ``.xlsx`` with the standard formatting.

    Used both for the blank table and for re-writing it after
    :func:`update_matchup_table` has filled the derivable columns.

    Args:
        df: A frame carrying exactly :data:`ALL_COLUMNS`.
        path_output: Destination ``.xlsx`` path.

    Returns:
        The written path.
    """
    path_output = Path(path_output)
    path_output.parent.mkdir(parents=True, exist_ok=True)
    df = df.reindex(columns=ALL_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "MatchupTable"

    # Row 1: the merged group banner.
    ws.append(["Network"] * len(NETWORK_COLUMNS)
              + ["Demand"] * len(DEMAND_COLUMNS)
              + ["Signal"] * len(SIGNAL_COLUMNS)
              + [""] * len(OTHER_COLUMNS)
              + [""] * len(EXTRA_COLUMNS))
    n_net, n_dem, n_sig = len(NETWORK_COLUMNS), len(DEMAND_COLUMNS), len(SIGNAL_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_net)
    ws.merge_cells(start_row=1, start_column=n_net + 1, end_row=1,
                   end_column=n_net + n_dem)
    ws.merge_cells(start_row=1, start_column=n_net + n_dem + 1, end_row=1,
                   end_column=n_net + n_dem + n_sig)

    # Row 2: the real header.  Data starts at row 3.
    ws.append(ALL_COLUMNS)
    for row in df.itertuples(index=False):
        ws.append([None if pd.isna(v) else v for v in row])

    _merge_junction_blocks(ws, df)

    # One UTDF export covers the network, so File_Synchro spans every data row.
    if len(df) > 1:
        col = ALL_COLUMNS.index("File_Synchro") + 1
        ws.merge_cells(start_row=3, start_column=col, end_row=len(df) + 2,
                       end_column=col)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A3"

    wb.save(path_output)
    return path_output


def _merge_junction_blocks(ws, df: pd.DataFrame) -> None:
    """Vertically merge the cells that are constant within a junction.

    Args:
        ws: The openpyxl worksheet, already populated.
        df: The dataframe that was written, used for the junction boundaries.
    """
    merge_cols = [ALL_COLUMNS.index(c) + 1 for c in MERGED_COLUMNS]
    junction_ids = df["JunctionID_Vissim"].tolist()

    start = 3  # data starts at row 3
    for i, junction_id in enumerate(junction_ids):
        row = i + 3
        is_last = i == len(junction_ids) - 1
        next_id = None if is_last else junction_ids[i + 1]
        if is_last or next_id != junction_id:
            if row > start:  # only merge genuine multi-row blocks
                for col in merge_cols:
                    ws.merge_cells(start_row=start, start_column=col,
                                   end_row=row, end_column=col)
            start = row + 1


# ---------------------------------------------------------------------- #
# Filling it in
# ---------------------------------------------------------------------- #
#: RealTwin turn label -> movement-code suffix.
TURN_LETTERS = {"right": "R", "thru": "T", "left": "L", "Uturn": "U"}

#: Bound codes in the order RealTwin walks a junction's legs.
BOUND_ORDER = ["NB", "EB", "SB", "WB"]

#: RealTwin orders a junction's legs clockwise from NNW rather than from north,
#: so a leg just west of north sorts first.  See ``format_junction_bearing``.
BEARING_ORIGIN = 337.5


def assign_movement_codes(group: pd.DataFrame, available: set[str]) -> pd.Series:
    """Map each row of one junction to a ``NBR``-style movement code.

    A bound cannot be read off a single approach's bearing.  Junction 8 in
    Chattanooga has legs at 52 deg and 114 deg, both inside any absolute
    "eastbound" sector, yet the count file calls them northbound and eastbound;
    its third leg at 297 deg is westbound with no southbound leg at all.  What
    fixes the bound is the leg's *rank* around the junction paired with the
    bounds the source file actually reports.

    So the legs are ordered the way RealTwin orders them -- clockwise from
    :data:`BEARING_ORIGIN` -- and the n-th leg takes the n-th bound present in
    ``available``.

    Args:
        group: The rows of one junction, carrying ``Bearing``,
            ``FromLinkNo_Vissim`` and ``Turn``.
        available: Movement codes the source file reports for this junction.

    Returns:
        Codes aligned to ``group.index``; ``None`` where the row's movement is
        not one the source reports.
    """
    bounds = [b for b in BOUND_ORDER if any(c.startswith(b) for c in available)]

    legs = group.copy()
    legs["_shift"] = (pd.to_numeric(legs["Bearing"], errors="coerce")
                      - BEARING_ORIGIN) % 360
    ordered = (legs.dropna(subset=["_shift"])
               .drop_duplicates(subset=["FromLinkNo_Vissim"])
               .sort_values("_shift")["FromLinkNo_Vissim"].tolist())
    leg_bound = dict(zip(ordered, bounds))

    codes = []
    for _, row in group.iterrows():
        bound = leg_bound.get(row["FromLinkNo_Vissim"])
        letter = TURN_LETTERS.get(str(row["Turn"]))
        code = f"{bound}{letter}" if bound and letter else None
        codes.append(code if code in available else None)
    series = pd.Series(codes, index=group.index)

    # Turn classification near the thru/turn threshold can label two movements
    # off one approach identically, which would join a single turn count onto
    # both and overstate the approach.  Movements are ordered by turn angle, so
    # their codes must be strictly increasing in R, T, L, U order; re-assign any
    # approach where that fails.
    for from_link, rows in group.groupby("FromLinkNo_Vissim", sort=False):
        filled = series.loc[rows.index].dropna()
        if filled.empty or not filled.duplicated().any():
            continue
        bound = leg_bound.get(from_link)
        if bound is None:
            continue
        options = [f"{bound}{letter}" for letter in "RTLU"
                   if f"{bound}{letter}" in available]
        resolved = _monotone_codes([series.get(i) for i in rows.index], options)
        for index, code in zip(rows.index, resolved):
            series.at[index] = code
    return series


def _monotone_codes(current: list, options: list[str]) -> list:
    """Re-assign one approach's codes so they increase through R, T, L, U.

    Picks the strictly increasing choice from ``options`` that departs least from
    the geometric labels, so a movement is only relabelled when the ordering
    forces it.

    Args:
        current: The geometric codes, in turn-angle order; may contain ``None``.
        options: The bound's available codes, in ``R``, ``T``, ``L``, ``U`` order.

    Returns:
        A list the same length as ``current``.
    """
    if not options or len(current) > len(options):
        return current

    rank = {code: i for i, code in enumerate(options)}
    best: tuple[int, list] | None = None

    def search(position: int, start: int, cost: int, chosen: list) -> None:
        nonlocal best
        if best is not None and cost >= best[0]:
            return
        if position == len(current):
            best = (cost, list(chosen))
            return
        for index in range(start, len(options)):
            # Enough options must remain for the movements still to be placed.
            if len(options) - index < len(current) - position:
                break
            want = rank.get(current[position])
            penalty = 0 if want is None else abs(index - want)
            chosen.append(options[index])
            search(position + 1, index + 1, cost + penalty, chosen)
            chosen.pop()

    search(0, 0, 0, [])
    return best[1] if best else current


def _gridsmart_info(path: Path) -> tuple[str | None, str | None, set[str]]:
    """Read a GridSmart export's header block.

    Args:
        path: Path to the ``.xls`` / ``.xlsx`` export.

    Returns:
        ``(intersection_name, date, movement_codes)``.  Movement codes are the
        ``NBR``-style codes the file actually reports; as in RealTwin, a
        reported left implies the matching U-turn.
    """
    raw = pd.read_excel(path, header=None, dtype=str)

    def _header_value(label: str) -> str | None:
        rows = raw[raw.iloc[:, 0] == label].index
        if rows.empty:
            return None
        col = raw.iloc[rows[0], 1:].first_valid_index()
        return None if col is None else raw.iloc[rows[0], col]

    # Locate the columns that carry each movement, then see which directions
    # actually report a value under them.
    movement_columns: dict[str, int] = {}
    for col in raw.columns[1:]:
        for movement in ("Right", "Through", "Left"):
            if raw.iloc[:, col].eq(movement).any():
                movement_columns[movement] = col

    codes: set[str] = set()
    for direction, prefix in (("Northbound", "NB"), ("Eastbound", "EB"),
                              ("Southbound", "SB"), ("Westbound", "WB")):
        rows = raw[raw.iloc[:, 0] == direction].index
        if rows.empty:
            continue
        row = rows[0]
        for movement, suffix in (("Right", "R"), ("Through", "T"), ("Left", "L")):
            col = movement_columns.get(movement)
            if col is not None and pd.notna(raw.iloc[row, col]):
                codes.add(f"{prefix}{suffix}")
                if suffix == "L":
                    codes.add(f"{prefix}U")
    return _header_value("Intersection"), _header_value("Date"), codes


def _synchro_movements(signal_dict: dict, intid: str) -> set[str]:
    """Return the movement codes one Synchro intersection serves.

    Mirrors the column-pruning rule in RealTwin's ``update_matchup_table``: a
    movement column survives when the ``Lanes`` row is non-zero, or a later
    phase row carries a value, and a reported left implies the U-turn.

    Args:
        signal_dict: Output of RealTwin's ``process_signal_from_utdf``.
        intid: The Synchro ``INTID`` to select.

    Returns:
        The movement codes, e.g. ``{"NBL", "NBT", "NBU", ...}``.
    """
    lanes = signal_dict.get("Lanes")
    if lanes is None:
        return set()

    allowed = ["Lanes", "Shared", "Phase1", "PermPhase1", "Phase2",
               "PermPhase2", "Phase3", "PermPhase3"]
    subset = lanes[(lanes["INTID"].astype(str) == str(intid))
                   & (lanes["RECORDNAME"].astype(str).isin(allowed))]
    if subset.empty:
        return set()
    subset = subset.reset_index(drop=True)

    def _blank(val) -> bool:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return True
        text = str(val).strip()
        if text in ("", "nan", "None"):
            return True
        try:
            return float(text) == 0
        except ValueError:
            return False

    def _row(name: str) -> pd.Series | None:
        hit = subset[subset["RECORDNAME"] == name]
        return None if hit.empty else hit.iloc[0]

    lanes_row, shared_row = _row("Lanes"), _row("Shared")
    phase_row, perm_row = _row("Phase1"), _row("PermPhase1")

    # A shared lane serves a turn that has no phase of its own: Synchro records
    # the through phase and a Shared code (1 left, 2 right, 3 both).  Without
    # this the turn looks unserved and its column is pruned below.
    shared_turns: set[str] = set()
    if shared_row is not None and phase_row is not None:
        for col in subset.columns:
            if not col.endswith("T"):
                continue
            if _blank(phase_row.get(col)) or _blank(shared_row.get(col)):
                continue
            try:
                share = int(float(shared_row[col]))
            except (TypeError, ValueError):
                continue
            base = col[:-1]
            if share in (1, 3) and f"{base}L" in subset.columns:
                shared_turns.add(f"{base}L")
            if share in (2, 3) and f"{base}R" in subset.columns:
                shared_turns.add(f"{base}R")

    codes: set[str] = set()
    for col in subset.columns:
        if col in ("RECORDNAME", "INTID", "PED", "HOLD") or len(col) < 3:
            continue
        if col[-1] not in ("R", "T", "L", "U"):
            continue

        keep = not _blank(subset.at[0, col])
        if not keep and len(subset) > 2:
            keep = any(not _blank(v) for v in subset[col].iloc[2:])
        if not keep:
            keep = col in shared_turns
        # A right turn off a multi-lane through movement is served even when it
        # has no lane count of its own.
        if not keep and col.endswith("R") and lanes_row is not None:
            through = f"{col[:-1]}T"
            if through in subset.columns and shared_row is not None:
                try:
                    keep = (float(lanes_row[through]) > 0
                            and float(shared_row[through]) > 1)
                except (TypeError, ValueError):
                    pass
        if keep:
            codes.add(col)
            if col.endswith("L"):
                codes.add(col[:-1] + "U")
    return codes


#: Movement-code suffix back to the RealTwin turn label.
LETTER_TURNS = {letter: turn for turn, letter in TURN_LETTERS.items()}


def _reconcile_turns(df: pd.DataFrame, group: pd.DataFrame,
                     assigned: pd.Series) -> int:
    """Write an ordering-derived turn back into the ``Turn`` column.

    ``assign_movement_codes`` relabels a movement when the geometric turn breaks
    the R, T, L, U ordering along an approach.  Correcting only the code leaves
    the table contradicting itself -- ``Turn`` saying ``left`` beside a code of
    ``NBT`` -- and leaves the wrong label in the column the routing stage reads.

    The ordering is the better evidence.  Two movements off one approach cannot
    both be lefts when they are sorted by turning angle, whereas the 20 degree
    thru/turn threshold is a single fixed cut that skewed approaches sit right on
    top of.

    Args:
        df: The full table, updated in place.
        group: The junction's rows.
        assigned: The codes just assigned, aligned to ``group``.

    Returns:
        How many turn labels were corrected.
    """
    corrected = 0
    for index, code in assigned.dropna().items():
        turn = LETTER_TURNS.get(str(code)[-1])
        current = str(group.at[index, "Turn"])
        if turn and turn != current:
            print(f"  :NOTE: movement {group.at[index, 'FromLinkNo_Vissim']}"
                  f"->{group.at[index, 'ToLinkNo_Vissim']} classified '{current}' by "
                  f"bearing, but the turn order at this approach makes it '{turn}'; "
                  "corrected.")
            df.at[index, "Turn"] = turn
            corrected += 1
    return corrected


def report_coverage(junction_id, column: str, group: pd.DataFrame,
                    available: set[str], assigned: pd.Series) -> None:
    """Warn when a junction's legs and the source's directions do not line up.

    The rank rule pairs the n-th leg with the n-th direction the source reports,
    which is only sound when there are as many of one as the other.  Two things
    can go wrong quietly:

    * more legs than directions, so the surplus legs silently take no code and
      their movements are never given a count;
    * a direction the source reports that no movement claims, which means the
      pairing is rotated -- every leg has taken its neighbour's name.

    Neither is recoverable here; the MatchupTable is the place to correct them,
    exactly as in the SUMO flow.  This makes them visible.

    Args:
        junction_id: The junction being filled.
        column: Which column is being filled, for the message.
        group: The junction's rows.
        available: Movement codes the source reports.
        assigned: The codes just assigned, aligned to ``group``.
    """
    bounds = [b for b in BOUND_ORDER if any(c.startswith(b) for c in available)]
    legs = group["FromLinkNo_Vissim"].nunique()
    if legs != len(bounds):
        print(f"  :WARNING: junction {junction_id} has {legs} approaches but the "
              f"source reports {len(bounds)} directions {bounds} for {column}; "
              f"{'surplus approaches get no code' if legs > len(bounds) else 'a direction is unused'}.")

    unused = sorted(available - set(assigned.dropna()))
    if unused:
        print(f"  :WARNING: junction {junction_id}: {column} codes {unused} are "
              "reported by the source but matched no movement, so those counts are "
              "dropped. Check the approach bearings, or set the codes by hand.")


def update_matchup_table(path_matchup_table: str | Path, *,
                         traffic_dir: str | Path = "",
                         control_dir: str | Path = "") -> pd.DataFrame:
    """Fill every derivable MatchupTable column in place.

    The user supplies only the per-junction seeds -- which GridSmart file and
    which Synchro ``INTID`` belong to each junction, plus the one ``File_Synchro``
    -- exactly as in RealTwin's ``update_matchup_table``.  Everything else is
    derived here: the intersection name and date from the GridSmart header, and
    the per-movement ``Turn_GridSmart`` / ``Turn_Synchro`` codes.

    Unlike the SUMO version, which assigns turn codes to rows *positionally* and
    so depends on the row order matching the file's movement order, this derives
    each row's code from its own ``Bearing`` and ``Turn`` and keeps it only if
    the source actually reports that movement.

    Args:
        path_matchup_table: The table to update, rewritten in place.
        traffic_dir: Directory holding the GridSmart files.
        control_dir: Directory holding the Synchro UTDF file.

    Returns:
        The filled frame.

    Raises:
        FileNotFoundError: If ``path_matchup_table`` does not exist.
    """
    path_matchup_table = Path(path_matchup_table)
    if not path_matchup_table.exists():
        raise FileNotFoundError(f"MatchupTable not found: {path_matchup_table}")

    df = pd.read_excel(path_matchup_table, skiprows=1, dtype=str)
    df = df.reindex(columns=ALL_COLUMNS)
    df[["JunctionID_Vissim", "File_Synchro"]] = (
        df[["JunctionID_Vissim", "File_Synchro"]].ffill().infer_objects(copy=False))
    df["Need calibration?"] = "Y"
    corrected = 0

    # -- demand ------------------------------------------------------- #
    for junction_id, group in df.groupby("JunctionID_Vissim", sort=False):
        files = group["File_GridSmart"].dropna()
        if files.empty:
            continue
        name = str(files.iloc[0])
        if "." not in Path(name).name:
            name += ".xls"
        path = Path(traffic_dir) / name
        if not path.exists():
            print(f"  :WARNING: GridSmart file missing for junction {junction_id}: {name}")
            continue

        try:
            intersection, date, codes = _gridsmart_info(path)
        except (ValueError, KeyError) as exc:
            print(f"  :WARNING: could not read {path.name}: {exc}")
            continue

        rows = df["JunctionID_Vissim"] == junction_id
        df.loc[rows, "Need calibration?"] = "N"
        if intersection:
            df.loc[rows, "IntersectionName_GridSmart"] = intersection
        if date:
            df.loc[rows, "Date_GridSmart"] = date
        assigned = assign_movement_codes(group, codes)
        df.loc[rows, "Turn_GridSmart"] = assigned
        report_coverage(junction_id, "Turn_GridSmart", group, codes, assigned)
        corrected += _reconcile_turns(df, group, assigned)

    # -- signal ------------------------------------------------------- #
    cache: dict[str, dict] = {}
    for junction_id, group in df.groupby("JunctionID_Vissim", sort=False):
        intids = group["IntersectionID_Synchro"].dropna()
        files = group["File_Synchro"].dropna()
        if intids.empty or files.empty:
            continue

        utdf = str(files.iloc[0])
        if utdf not in cache:
            try:
                from realtwin.func_lib._c_abstract_scenario.rt_demand_generation import (
                    process_signal_from_utdf)
                cache[utdf] = process_signal_from_utdf(str(Path(control_dir) / utdf))
            except (ImportError, FileNotFoundError, ValueError, KeyError) as exc:
                print(f"  :WARNING: could not read Synchro file {utdf}: {exc}")
                cache[utdf] = {}
        codes = _synchro_movements(cache[utdf], str(intids.iloc[0]))
        if not codes:
            continue
        rows = df["JunctionID_Vissim"] == junction_id
        assigned = assign_movement_codes(group, codes)
        df.loc[rows, "Turn_Synchro"] = assigned
        report_coverage(junction_id, "Turn_Synchro", group, codes, assigned)

    # A movement code must be unique within a junction: two rows claiming the
    # same code would join the same turn count twice.  Duplicates mean the
    # bearing or turn classification disagrees with the count file, so flag the
    # junction for review rather than emitting a table that quietly double counts.
    for col in ("Turn_GridSmart", "Turn_Synchro"):
        for junction_id, group in df.groupby("JunctionID_Vissim", sort=False):
            filled = group[col].dropna()
            dupes = sorted(set(filled[filled.duplicated()]))
            if dupes:
                print(f"  :WARNING: junction {junction_id} derives {col} "
                      f"{dupes} more than once; check the approach bearings. "
                      "Marked for calibration.")
                df.loc[df["JunctionID_Vissim"] == junction_id, "Need calibration?"] = "Y"

    if corrected:
        print(f"  :{corrected} turn labels corrected from the approach's turn order; "
              "the movement table from stage 1 still holds the bearing-derived value.")

    write_matchup_table(df, path_matchup_table)
    return df


# ---------------------------------------------------------------------- #
# Reading it back
# ---------------------------------------------------------------------- #
class MatchupTable:
    """A user-filled Vissim MatchupTable, normalised for downstream use.

    Args:
        path: Path to ``MatchupTable.xlsx``.

    Raises:
        FileNotFoundError: If ``path`` does not exist.
        ValueError: If ``path`` is not ``.xlsx`` or lacks the network columns.
    """

    def __init__(self, path: str | Path):
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"MatchupTable not found: {self.path}")
        if self.path.suffix.lower() != ".xlsx":
            raise ValueError(f"MatchupTable must be .xlsx, got: {self.path}")

        # Row 0 is the group banner; the real header is row 1.
        df = pd.read_excel(self.path, skiprows=1, dtype=str)

        missing = [c for c in NETWORK_COLUMNS if c not in df.columns]
        if missing:
            raise ValueError(f"MatchupTable {self.path.name} is missing columns: {missing}")

        for col in FILLED_COLUMNS:
            if col not in df.columns:
                df[col] = pd.NA

        # A merged cell stores its value only in the top-left cell, so the rest
        # of the block reads back empty and has to be filled.  JunctionID and
        # File_Synchro span the sheet; the others are merged per junction, so
        # filling them globally would leak one junction's count file onto the
        # next junction, which has none.
        # infer_objects avoids pandas' deprecated silent downcast on object ffill.
        sheet_wide = ["JunctionID_Vissim", "File_Synchro"]
        df[sheet_wide] = df[sheet_wide].ffill().infer_objects(copy=False)
        per_junction = [c for c in FILLED_COLUMNS if c not in sheet_wide]
        df[per_junction] = (df.groupby("JunctionID_Vissim")[per_junction]
                            .ffill().infer_objects(copy=False))

        # Hand-edited spreadsheets are full of stray whitespace.
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].astype(str).str.strip().replace({"nan": None, "": None})

        self.df = df

    # -- signal ------------------------------------------------------- #
    def synchro_file(self) -> str:
        """Return the Synchro UTDF filename referenced by the table.

        Raises:
            ValueError: If no ``File_Synchro`` value is present.
        """
        vals = [v for v in self.df["File_Synchro"].dropna().unique() if str(v).strip()]
        if not vals:
            raise ValueError(f"No File_Synchro value in {self.path.name}")
        if len(vals) > 1:
            print(f"  :NOTE: multiple File_Synchro values {vals}; using '{vals[0]}'.")
        return str(vals[0])

    def signalized_junctions(self) -> dict[str, str]:
        """Return ``{Vissim junction ID: Synchro INTID}`` for signalised junctions.

        A junction counts as signalised when at least one of its rows carries a
        ``Turn_Synchro`` code -- the rule RealTwin's ``parse_SUMO_TLS_ID`` uses
        to decide which junctions get a signal.
        """
        rows = self.df[self.df["Turn_Synchro"].notna()]
        out: dict[str, str] = {}
        for junction_id, group in rows.groupby("JunctionID_Vissim"):
            intids = [v for v in group["IntersectionID_Synchro"].dropna().unique() if str(v).strip()]
            if not intids:
                print(f"  :WARNING: junction {junction_id} has Turn_Synchro codes but no "
                      "IntersectionID_Synchro; skipping its signal.")
                continue
            out[str(junction_id)] = str(intids[0])
        return out

    # -- demand ------------------------------------------------------- #
    def gridsmart_files(self) -> dict[str, str]:
        """Return ``{Vissim junction ID: GridSmart turn-count filename}``."""
        out: dict[str, str] = {}
        for junction_id, group in self.df.groupby("JunctionID_Vissim"):
            files = [v for v in group["File_GridSmart"].dropna().unique() if str(v).strip()]
            if files:
                out[str(junction_id)] = str(files[0])
        return out

    def turn_lookup(self) -> pd.DataFrame:
        """Return the demand join table: GridSmart turn code to Vissim links.

        Returns:
            Columns ``IntersectionName``, ``Turn``, ``JunctionID_Vissim``,
            ``FromLinkNo_Vissim``, ``ToLinkNo_Vissim`` -- one row per movement
            that carries a GridSmart turn code.  This is the Vissim analogue of
            the ``IDRef`` frame RealTwin builds in ``generate_turn_demand_cali``.
        """
        rows = self.df[self.df["Turn_GridSmart"].notna()].copy()
        out = rows[["IntersectionName_GridSmart", "Turn_GridSmart", "JunctionID_Vissim",
                    "FromLinkNo_Vissim", "ToLinkNo_Vissim"]].copy()
        out.columns = ["IntersectionName", "Turn", "JunctionID_Vissim",
                       "FromLinkNo_Vissim", "ToLinkNo_Vissim"]
        out = out.dropna(subset=["FromLinkNo_Vissim", "ToLinkNo_Vissim"])
        for col in ("FromLinkNo_Vissim", "ToLinkNo_Vissim"):
            out[col] = out[col].astype(float).astype(int)
        return out.reset_index(drop=True)

    def link_numbers(self) -> set[int]:
        """Return every Vissim link number referenced by the table."""
        nos: set[int] = set()
        for col in ("FromLinkNo_Vissim", "ToLinkNo_Vissim"):
            for v in self.df[col].dropna():
                try:
                    nos.add(int(float(v)))
                except (TypeError, ValueError):
                    continue
        return nos

    def __repr__(self) -> str:  # pragma: no cover - debugging aid
        return (f"MatchupTable(path={self.path.name!r}, rows={len(self.df)}, "
                f"junctions={self.df['JunctionID_Vissim'].nunique()})")
