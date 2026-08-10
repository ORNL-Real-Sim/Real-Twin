##############################################################################
# Copyright (c) 2024-, Oak Ridge National Laboratory                          #
# All rights reserved.                                                       #
#                                                                            #
# This file is part of RealTwin and is distributed under a GPL               #
# license. For the licensing terms see the LICENSE file in the top-level     #
# directory.                                                                 #
#                                                                            #
# Contributors: ORNL Real-Twin Team                                          #
# Contact: realtwin@ornl.gov                                                 #
##############################################################################
"""Stage 2: build the Vissim MatchupTable and fill everything derivable.

Takes the movement table stage 1 wrote and produces ``MatchupTable.xlsx``.

The table needs one small piece of human input: which GridSmart file and which
Synchro ``INTID`` belong to each junction, plus the single ``File_Synchro``.  On
Chattanooga that is 13 cells out of 104 rows.  Everything else -- the
intersection name, the date, ``Need calibration?`` and all 80 per-movement
``Turn_GridSmart`` / ``Turn_Synchro`` codes -- is derived by
:func:`rt_vissim.matchup.update_matchup_table`.

``--seed-from`` lifts those seeds out of an existing SUMO MatchupTable, which is
how the Vissim table is checked against the SUMO one on the same network.

Usage::

    python vissim/scripts/02_matchup_table.py
    python vissim/scripts/02_matchup_table.py --seed-from datasets/chattanooga/updated_net/MatchupTable.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
# The repository root, so the Synchro UTDF parser in realtwin is importable.
sys.path.insert(1, str(Path(__file__).resolve().parents[2]))

from rt_vissim.demand import build_turn_counts  # noqa: E402
from rt_vissim.matchup import (  # noqa: E402
    ALL_COLUMNS, MatchupTable, generate_matchup_table, update_matchup_table)
from rt_vissim.validate import validate  # noqa: E402

#: Columns a user seeds by hand, and where they live in the sheet.
SEED_COLUMNS = {"File_GridSmart": ALL_COLUMNS.index("File_GridSmart") + 1,
                "IntersectionID_Synchro": ALL_COLUMNS.index("IntersectionID_Synchro") + 1}


def seed_from_sumo(path_table: Path, path_sumo: Path,
                   path_junctions: Path | None = None) -> int:
    """Copy the per-junction seeds out of a SUMO MatchupTable.

    The two tables do **not** share junction IDs.  The Vissim table is keyed on
    OpenDRIVE junction IDs, the SUMO one on SUMO's own, and on Chattanooga they
    overlap enough to look right while being wrong: OpenDRIVE 10 is SUMO 8, so
    seeding by the bare number hands junction 10 the count file for a different
    intersection and says nothing.

    So the seeds are translated through the mapping stage 1 recorded in
    ``<name>_junctions.csv``, whose ``Name_OpenDrive`` column holds the SUMO
    junction ID that ``netconvert`` wrote into the OpenDRIVE junction's name.

    Args:
        path_table: The Vissim table to seed, edited in place.
        path_sumo: A filled SUMO ``MatchupTable.xlsx``.
        path_junctions: ``<name>_junctions.csv`` from stage 1.  Without it the
            IDs are assumed to coincide, which is only safe when the SUMO table
            was itself keyed on OpenDRIVE IDs.

    Returns:
        How many junctions were seeded.
    """
    sumo = pd.read_excel(path_sumo, skiprows=1)
    sumo["JunctionID_OpenDrive"] = sumo["JunctionID_OpenDrive"].ffill()
    seeds = sumo.groupby("JunctionID_OpenDrive").agg(
        File_GridSmart=("File_GridSmart", "first"),
        IntersectionID_Synchro=("IntersectionID_Synchro", "first"))
    utdf = sumo["File_Synchro"].dropna()

    # OpenDRIVE junction ID -> the key the SUMO table is written against.
    translate: dict[float, float] = {}
    if path_junctions is not None and Path(path_junctions).exists():
        mapping = pd.read_csv(path_junctions)
        for row in mapping.itertuples(index=False):
            try:
                translate[float(row.JunctionID_OpenDrive)] = float(row.Name_OpenDrive)
            except (TypeError, ValueError):
                continue  # a "J"-prefixed SUMO node, never a real intersection
        if translate:
            print(f"  :Translating {len(translate)} junction IDs through "
                  f"{Path(path_junctions).name}")

    wb = load_workbook(path_table)
    ws = wb.active

    # A junction's merged cells hold their value in the block's first row.
    first_row: dict[float, int] = {}
    for row in range(3, ws.max_row + 1):
        junction = ws.cell(row, 1).value
        if junction is not None:
            first_row.setdefault(float(junction), row)

    seeded = 0
    missing = []
    for junction, row in first_row.items():
        key = translate.get(junction, junction)
        if key not in seeds.index:
            missing.append(junction)
            continue
        used = False
        for col_name, col in SEED_COLUMNS.items():
            value = seeds.loc[key, col_name]
            if pd.isna(value):
                continue
            ws.cell(row, col).value = (str(int(value))
                                       if col_name.endswith("_Synchro") else str(value))
            used = True
        seeded += used

    if missing:
        print(f"  :WARNING: {len(missing)} junctions found no seed in "
              f"{Path(path_sumo).name}: {sorted(missing)}. Their demand and "
              "signal columns stay blank for you to fill in.")

    if not utdf.empty:
        ws.cell(3, ALL_COLUMNS.index("File_Synchro") + 1).value = str(utdf.iloc[0])
    wb.save(path_table)
    return seeded


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--movements", default="vissim/work/chattanooga/chatt_movements.csv",
                        help="Movement table written by stage 1")
    parser.add_argument("--outdir", default="vissim/work/chattanooga",
                        help="Where to write MatchupTable.xlsx")
    parser.add_argument("--traffic-dir", default="datasets/chattanooga/Traffic",
                        help="Directory holding the GridSmart files")
    parser.add_argument("--control-dir", default="datasets/chattanooga/Control",
                        help="Directory holding the Synchro UTDF file")
    parser.add_argument("--seed-from", default=None,
                        help="A SUMO MatchupTable to copy the per-junction seeds from")
    parser.add_argument("--no-update", action="store_true",
                        help="Write the blank table only, skip the auto-fill")
    parser.add_argument("--regenerate", action="store_true",
                        help="Rebuild the filled columns, discarding hand edits "
                             "(use after changing the derivation itself)")
    args = parser.parse_args(argv)

    movements_path = Path(args.movements)
    if not movements_path.exists():
        print(f"  :Movement table not found: {movements_path}\n"
              "  :Run 01_import_opendrive.py first.")
        return 1

    movements = pd.read_csv(movements_path)
    out_path = Path(args.outdir) / "MatchupTable.xlsx"
    generate_matchup_table(movements, out_path)
    print(f"  :Wrote {out_path} -- {len(movements)} movements, "
          f"{movements['JunctionID_Vissim'].nunique()} junctions")

    if args.seed_from:
        # Stage 1 writes the junction mapping beside the movement table.  Seeding
        # without it matches junction IDs across two different numbering spaces,
        # which hands a junction another intersection's count file.
        junctions_csv = movements_path.with_name(
            movements_path.name.replace("_movements.csv", "_junctions.csv"))
        seeded = seed_from_sumo(out_path, Path(args.seed_from), junctions_csv)
        print(f"  :Seeded {seeded} junctions from {Path(args.seed_from).name}")
    elif not args.no_update:
        print("  :No --seed-from given; fill File_GridSmart, IntersectionID_Synchro "
              "and File_Synchro by hand, then re-run to derive the rest.")

    if args.no_update:
        return 0

    df = update_matchup_table(out_path, traffic_dir=args.traffic_dir,
                              control_dir=args.control_dir,
                              preserve_edits=not args.regenerate)
    for col in ("IntersectionName_GridSmart", "Turn_GridSmart", "Turn_Synchro"):
        print(f"  :{col:<28} filled {df[col].notna().sum():>3}/{len(df)}")

    table = MatchupTable(out_path)
    print(f"  :{len(table.gridsmart_files())} junctions with counts, "
          f"{len(table.signalized_junctions())} signalised, "
          f"{len(table.turn_lookup())} movements joined to a turn count")

    # Reference-free checks, so the table can be trusted on a corridor that has
    # no SUMO MatchupTable to compare against.
    turn_counts = build_turn_counts(table, args.traffic_dir)
    findings = validate(movements, table.df, turn_counts)
    print(f"  :Validation: {len(findings)} findings"
          + (f" ({sum(1 for f in findings if f.severity == 'error')} errors)"
             if findings else " - nothing to report"))
    for finding in findings:
        print(finding)
    if findings:
        print("  :Correct these in the MatchupTable itself; a re-run keeps what you "
              "enter unless --regenerate is given.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
