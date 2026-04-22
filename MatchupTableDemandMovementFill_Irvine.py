import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import os


# Please refer to the official documentation for more details on RealTwin preparation before running the simulation
# # https://real-twin.readthedocs.io/en/latest/index.html
# add a function to modify matchup_table by filling out the demand movement column
def generate_matchup_excel(MatchupTable_OpenDrive, output_filename="MatchupTable_OpenDrive.xlsx"):
    network_columns = ["JunctionID_OpenDrive", "Bearing", "Numbering", "FromRoadID_OpenDrive", "ToRoadID_OpenDrive",
                       "Turn"]
    demand_columns = ["File_GridSmart", "Date_GridSmart", "IntersectionName_GridSmart", "Turn_GridSmart"]
    signal_columns = ["File_Synchro", "IntersectionID_Synchro", "Turn_Synchro"]
    other_columns = ["Need calibration?"]

    wb = Workbook()
    ws = wb.active

    ws.append(["Network"] * len(network_columns) + ["Demand"] * len(demand_columns) +
              ["Signal"] * len(signal_columns) + [""] * len(other_columns))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(network_columns))
    ws.merge_cells(start_row=1, start_column=len(network_columns) + 1, end_row=1,
                   end_column=len(network_columns) + len(demand_columns))
    ws.merge_cells(start_row=1, start_column=len(network_columns) + len(demand_columns) + 1, end_row=1,
                   end_column=len(network_columns) + len(demand_columns) + len(signal_columns))

    ws.append(network_columns + demand_columns + signal_columns + other_columns)

    for row in MatchupTable_OpenDrive.itertuples(index=False):
        ws.append(list(row))

    current_start = 3  # Data starts at row 3
    for i in range(3, len(MatchupTable_OpenDrive) + 3):
        if (i == len(MatchupTable_OpenDrive) + 2 or
                ws[f"A{i}"].value != ws[f"A{i + 1}"].value):  # Check next row
            if current_start < i:  # Only merge if there are multiple same values
                ws.merge_cells(start_row=current_start, start_column=1, end_row=i, end_column=1)  # JunctionID_OpenDrive
                ws.merge_cells(start_row=current_start, start_column=7, end_row=i, end_column=7)  # File_GridSmart
                ws.merge_cells(start_row=current_start, start_column=8, end_row=i, end_column=8)  # Date_GridSmart
                ws.merge_cells(start_row=current_start, start_column=9, end_row=i,
                               end_column=9)  # IntersectionName_GridSmart
                # ws.merge_cells(start_row=current_start, start_column=11, end_row=i, end_column=11)  # File_Synchro
                ws.merge_cells(start_row=current_start, start_column=12, end_row=i,
                               end_column=12)  # IntersectionID_Synchro
                ws.merge_cells(start_row=current_start, start_column=14, end_row=i, end_column=14)  # Need calibration?
            current_start = i + 1

    if len(MatchupTable_OpenDrive) > 0:
        ws.merge_cells(start_row=3, start_column=11, end_row=len(MatchupTable_OpenDrive) + 2, end_column=11)

    # Center align merged cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    column_widths = {
        "A": 20,  # JunctionID_OpenDrive
        "B": 15,  # Bearing
        "C": 15,  # Numbering
        "D": 25,  # FromRoadID_OpenDrive
        "E": 25,  # ToRoadID_OpenDrive
        "F": 15,  # Turn
        "G": 20,  # File_GridSmart
        "H": 20,  # Date_GridSmart
        "I": 30,  # IntersectionName_GridSmart
        "J": 20,  # Turn_GridSmart
        "K": 20,  # File_Synchro
        "L": 25,  # IntersectionID_Synchro
        "M": 20,  # Turn_Synchro
        "N": 20  # Need calibration?
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_filename)


def determine_direction(angle: float) -> str:
    """
    Convert an angle into a compass direction.
    0° = North, positive = clockwise.

    Parameters:
    angle (float): Angle in degrees, ideally between -360 and 360.

    Returns:
    str: Compass direction label (NB, NE, EB, etc.), or None if invalid.
    """
    if angle is None or np.isnan(angle):
        return None

    if -45 <= angle <= 45 or angle >= 315 or angle <= -315:
        return 'NB'
    elif 45 < angle < 135 or -315 < angle < -225:
        return 'EB'
    elif 135 <= angle <= 225 or -225 <= angle <= -135:
        return 'SB'
    elif -135 < angle < -45 or 225 < angle < 315:
        return 'WB'

    return None


def fix_group(g: pd.DataFrame,
              name_col='approach_direction_bound',
              rank_col='approach_name_rank',
              approach_seq = ['NB', 'EB', 'SB', 'WB']):
    g = g.copy()

    seq_index = {v: i for i, v in enumerate(approach_seq)}
    # We’ll maintain a dynamic set of what's present as we modify labels
    present = set(g[name_col].tolist())

    counts = g[name_col].value_counts()
    dups_labels = [lab for lab, cnt in counts.items() if cnt > 1]

    if len(dups_labels) > 0:
        for lab in dups_labels:
            idxs = (
                g[g[name_col] == lab]
                .sort_values(rank_col, kind='stable')
                .index.tolist()
            )

            angles = (
                g[g[name_col] == lab]
                .sort_values(rank_col, kind='stable')
                .Bearing_int.tolist()
            )
            if len(idxs) < 2:
                continue

            first_idx, second_idx, *rest = idxs
            first_angle, second_angle, *rest = angles

            angle_diff = second_angle - first_angle

            li = seq_index[lab]
            left  = approach_seq[(li - 1) % 4]   # CCW neighbor
            right = approach_seq[(li + 1) % 4]   # CW neighbor

            # --- NEW: neighbor existence checks (no global 'missing') ---
            left_missing  = left  not in present
            right_missing = right not in present

            if lab == 'NB' and angle_diff > 45:
                # NB is a special case, [315 to 360] and [0 to 45]
                if left_missing and right_missing:
                    # both neighbors missing → replace only the FIRST with left
                    g.at[second_idx, name_col] = left

                else:
                    if left_missing:
                        g.at[second_idx, name_col] = left

                    elif right_missing:
                        g.at[first_idx, name_col] = right
                    else:
                        fixed_label = approach_seq[0:len(g)]
                        g[name_col] = fixed_label
            else:
                # Apply your rule:
                if left_missing and right_missing:
                    # both neighbors missing → replace only the FIRST with left
                    g.at[first_idx, name_col] = left

                else:
                    # If left is missing -> replace FIRST occurrence with left
                    if left_missing:
                        g.at[first_idx, name_col] = left

                    # If right is missing -> replace SECOND occurrence with right
                    elif right_missing:
                        g.at[second_idx, name_col] = right

                    # none of left and right is missing. fix the output
                    # if they are 4 approaches, we need to ensure they are all filled
                    # if they are more than 4 approaches, we need to label the second occurrence as the A + "_1"
                    elif len(g) > 4:
                        g.at[second_idx, name_col] = right + "_1"
                    else:
                        fixed_label = approach_seq[0:len(g)]
                        g[name_col] = fixed_label
        return g
    else:
        return g


if __name__ == '__main__':

    # set the project folder path
    project_folder_path = "datasets/Irvine"
    matchup_file_path = f"{project_folder_path}/MatchupTable.xlsx"

    MatchupTable_UserInput = pd.read_excel(matchup_file_path, skiprows=1, dtype=str)
    merged_columns = ["File_GridSmart", "Date_GridSmart", "IntersectionName_GridSmart",
                      "File_Synchro", "IntersectionID_Synchro", "Need calibration?"]
    
    # group-based forward fill merged_columns based on JunctionID_OpenDrive
    MatchupTable_UserInput["JunctionID_OpenDrive"] = MatchupTable_UserInput["JunctionID_OpenDrive"].ffill()
    # Method 1: Using transform with ffill
    for col in merged_columns:
        MatchupTable_UserInput[col] = MatchupTable_UserInput.groupby('JunctionID_OpenDrive')[col].transform('ffill')

    MatchupTable_UserInput['Numbering'] = MatchupTable_UserInput['Numbering'].astype(int)
    MatchupTable_UserInput['Bearing_int'] = MatchupTable_UserInput['Numbering'] * 10
    MatchupTable_UserInput['approach_direction_bound'] = MatchupTable_UserInput['Bearing_int'].apply(determine_direction)
    MatchupTable_UserInput = MatchupTable_UserInput.sort_values(by=['JunctionID_OpenDrive', 'Numbering']).reset_index(drop=True)

    df_approach_config = MatchupTable_UserInput.groupby(by=['JunctionID_OpenDrive', 'Numbering', 'Bearing_int', 'FromRoadID_OpenDrive'], as_index=False).agg(
        {'approach_direction_bound': 'first'})
    df_approach_config["approach_name_rank"] = df_approach_config.groupby(
        by=['JunctionID_OpenDrive', 'approach_direction_bound']).cumcount()

    # list all the junction ids in the folder datasets/Nashville1/Traffic
    traffic_folder_path = f"{project_folder_path}/Traffic"
    # get all the file names of the csv files in the folder
    traffic_files = [f.name for f in os.scandir(traffic_folder_path) if f.is_file() and f.name.endswith('.xls')]
    junction_ids_in_traffic = [i.split('.')[0] for i in traffic_files]

    # df_approach_config = df_approach_config[df_approach_config['JunctionID_OpenDrive'].isin(junction_ids_in_traffic)]
    df_approach_config = df_approach_config.groupby("JunctionID_OpenDrive", group_keys=False).apply(fix_group)

    # double check if there is still approach name overlap
    df_approach_config["approach_name_rank"] = df_approach_config.groupby(
        by=['JunctionID_OpenDrive', 'approach_direction_bound']).cumcount()

    # join df_approach_config back to MatchupTable_UserInput to update the approach_direction_bound
    MatchupTable_UserInput = MatchupTable_UserInput.merge(
        df_approach_config[['JunctionID_OpenDrive', 'Numbering', 'approach_direction_bound']],
        on=['JunctionID_OpenDrive', 'Numbering'],
        how='left',
        suffixes=('', '_updated')
    )

    # map the "Turn" column based on a dictionary. {'right': 'R', 'left': 'L', 'through': 'T', 'u-turn': 'U'}
    turn_mapping = {'right': 'R', 'left': 'L', 'thru': 'T', 'Uturn': 'U'}
    MatchupTable_UserInput['Turn_Label'] = MatchupTable_UserInput['Turn'].map(turn_mapping)

    MatchupTable_UserInput['Turn_GridSmart'] = MatchupTable_UserInput['approach_direction_bound_updated'] + MatchupTable_UserInput['Turn_Label']

    # remove duplicated rows in MatchupTable_UserInput
    MatchupTable_UserInput = MatchupTable_UserInput.drop_duplicates()

    ###################################################################################################
    # update the Turn_GridSmart based on RightTurnChannel_Junction_Movement_Config.csv  ###############
    ###################################################################################################
    # right_turn_mapping = pd.read_csv("datasets/Nashville1/RightTurnChannel_Junction_Movement_Config_0122.csv", dtype=str)


    # MatchupTable_UserInput = MatchupTable_UserInput.merge(
    #     right_turn_mapping[['From (Junction ID)', 'FromRoadID_OpenDrive', 'ToRoadID_OpenDrive', 'Turn_GridSmart']],
    #     left_on=['JunctionID_OpenDrive', 'FromRoadID_OpenDrive', 'ToRoadID_OpenDrive'],
    #     right_on=['From (Junction ID)', 'FromRoadID_OpenDrive', 'ToRoadID_OpenDrive'],
    #     how='left',
    #     suffixes=('', '_updated')
    # )
    # MatchupTable_UserInput['Turn_GridSmart'] = np.where(
    #     MatchupTable_UserInput['Turn_GridSmart_updated'].notnull(),
    #     MatchupTable_UserInput['Turn_GridSmart_updated'],
    #     MatchupTable_UserInput['Turn_GridSmart'])
    # MatchupTable_UserInput = MatchupTable_UserInput.drop(columns=['From (Junction ID)', 'Turn_GridSmart_updated'])

    ###################################################################################################
    # End section  ####################################################################################
    ###################################################################################################

    # join the signal_id based on another table sumo_network2signal_db_mapping.csv (sumo_id and signal_id)
    # signal_mapping = pd.read_csv(f"{project_folder_path}/sumo_network2signal_roosevelt_0227.csv", dtype=str)
    #
    # MatchupTable_UserInput = MatchupTable_UserInput.merge(
    #     signal_mapping[['sumo_id', 'signal_id']].drop_duplicates(),
    #     left_on='JunctionID_OpenDrive',
    #     right_on='sumo_id',
    #     how='left'
    # )

    # delete those movement that do not have GS data
    MatchupTable_UserInput['Turn_GridSmart'] = np.where(MatchupTable_UserInput['JunctionID_OpenDrive'].isin(junction_ids_in_traffic), MatchupTable_UserInput['Turn_GridSmart'], np.nan)

    MatchupTable_UserInput.to_csv(f"{project_folder_path}/MatchupTable_SUMO_Movements_Full_0212.csv", index=False)

    # update the column of "Need calibration?" based on whether signal_id is not null and Turn_GridSmart is null
    # MatchupTable_UserInput['Need calibration?'] = np.where(
    #     (MatchupTable_UserInput['signal_id'].notna()) &
    #     ((MatchupTable_UserInput['Turn_GridSmart'].isna()) | (~MatchupTable_UserInput['JunctionID_OpenDrive'].isin(junction_ids_in_traffic))),
    #     'Y',
    #     MatchupTable_UserInput['Need calibration?']
    # )
    ###################################################################################################

    # hard code the Turn_Synchro to address the issues that
    MatchupTable_UserInput.loc[(MatchupTable_UserInput['JunctionID_OpenDrive'] == '1')
                                & (MatchupTable_UserInput['FromRoadID_OpenDrive'] == '111563727#0')
                                & (MatchupTable_UserInput['ToRoadID_OpenDrive'] == '183526664'), 'Turn_Synchro'] = 'WBR'

    # group by junction and get the list of synchro approach name list for each junction, the synchro approach name is the first two digit of Turn_Synchro, e.g., if Turn_Synchro is SBT, then the approach name is SB.
    synchro_approach_map = MatchupTable_UserInput.groupby('JunctionID_OpenDrive')['Turn_Synchro'].apply(
        lambda x: list(dict.fromkeys(str(i)[:2] for i in x if pd.notna(i)))
    )
    MatchupTable_UserInput['synchro_approach_name_list'] = MatchupTable_UserInput['JunctionID_OpenDrive'].map(synchro_approach_map)

    # select the best synchro approach name from the synchro_approach_name_list based on the Bearing angle 
    # (0 to 360 degree, where 0 degree is north, 90 degree is east, 180 degree is south, 270 degree is west). 
    # the value in synchro_approach_name_list is one of NB, NE, EB, SE, SB, SW, WB, NW. We will select the one that has the smallest angle difference with the Bearing angle.
    # For example, if the Bearing is 350, synchro_approach_name_list is ['NB', 'SB'], then we will select NB as the best synchro approach name, if the Bearing is 100, synchro_approach_name_list is ['EB', 'WB'], then we will select EB as the best synchro approach name. 
    # We will create a new column "best_synchro_approach_name" to store the result.
    def get_angle_diff(bearing, approach):
        if approach == 'NB':
            return min(abs(bearing - 0), abs(bearing - 360))
        if approach == 'NE':
            return abs(bearing - 45)
        if approach == 'EB':
            return abs(bearing - 90)
        if approach == 'SE':
            return abs(bearing - 135)
        if approach == 'SB':
            return abs(bearing - 180)
        if approach == 'SW':
            return abs(bearing - 225)
        if approach == 'WB':
            return abs(bearing - 270)
        if approach == 'NW':
            return abs(bearing - 315)
        return np.inf

    def assign_unique_best_synchro_approach(group):
        approach_names = group['synchro_approach_name_list'].iloc[0]
        bearings = pd.to_numeric(group['Bearing'], errors='coerce').mod(360)

        if not isinstance(approach_names, list) or len(approach_names) == 0:
            return pd.Series(np.nan, index=group.index)

        unique_bearings = pd.unique(bearings.dropna())
        if len(unique_bearings) == 0:
            return pd.Series(np.nan, index=group.index)

        first_bearing = unique_bearings[0]
        first_label_idx = min(
            range(len(approach_names)),
            key=lambda idx: get_angle_diff(first_bearing, approach_names[idx])
        )

        bearing_to_approach = {}
        label_count = len(approach_names)

        for i, bearing in enumerate(unique_bearings):
            bearing_to_approach[bearing] = approach_names[(first_label_idx + i) % label_count]

        return bearings.map(bearing_to_approach)

    MatchupTable_UserInput['best_synchro_approach_name'] = (
        MatchupTable_UserInput.groupby('JunctionID_OpenDrive', group_keys=False)
        .apply(assign_unique_best_synchro_approach)
        .reset_index(level=0, drop=True)
    )

    MatchupTable_UserInput['Turn_Synchro'] = MatchupTable_UserInput['best_synchro_approach_name'] + MatchupTable_UserInput['Turn_Label']

    ###################################################################################################

    MatchupTable_UserInput['Need calibration?'] = np.where(MatchupTable_UserInput['Turn_GridSmart'].isna(),
        'Y', 'N')

    # verify if Turn_GridSmart is consistent with Turn_Synchro in terms of the turn type (R, L, T, U), if Turn_Label (R, L, T, U) in in Turn_Synchro (str, e.g., SBT), then it is consistent, otherwise it is not consistent and needs calibration. Create a new column "Synchro_GS_Consist_Verify" to store the result, if it is consistent, then the value is 0, otherwise the value is 1.
    # or if Turn_Synchro is 'NaN', then we also consider it as consistent, 0.
    MatchupTable_UserInput['Synchro_GS_Consist_Verify'] = np.where(
        MatchupTable_UserInput.apply(lambda row: any(label in str(row['Turn_Synchro']) for label in row['Turn_Label']), axis=1) | (~MatchupTable_UserInput['Turn_Synchro'].notnull()),
        0, 1)
    
    # groupby JunctionID_OpenDrive and calculate the sum of Synchro_GS_Consist_Verify, also keep the column of Turn_GridSmart and Turn_Synchro for reference, and save the result as consistency_summary
    consistency_summary = MatchupTable_UserInput.groupby('JunctionID_OpenDrive').agg(
        Synchro_GS_Consist_Sum=('Synchro_GS_Consist_Verify', 'sum'),
        Turn_GridSmart_List=('Turn_GridSmart', lambda x: list(x.dropna().unique())),
        Turn_Synchro_List=('Turn_Synchro', lambda x: list(x.dropna().unique()))
    ).reset_index()

    # save the consistency_summary to a csv file for future reference
    consistency_summary.to_csv(f"{project_folder_path}/Synchro_GS_Consistency_Summary.csv", index=False)

    # print the total number of inconsistent junctions, and the list of junction ids that are inconsistent
    inconsistent_junctions = consistency_summary[consistency_summary['Synchro_GS_Consist_Sum'] > 0]['JunctionID_OpenDrive'].tolist()
    print(f"Total number of inconsistent junctions: {len(inconsistent_junctions)}")
    print(f"List of inconsistent junction ids: {inconsistent_junctions}")

    # drop the temporary columns
    MatchupTable_UserInput.drop(columns=['Bearing_int', 'approach_direction_bound', 'approach_direction_bound_updated',
                                         'Turn_Label', 'synchro_approach_name_list',
                                         'best_synchro_approach_name', 'Synchro_GS_Consist_Verify'], inplace=True)
    generate_matchup_excel(MatchupTable_UserInput, f"{project_folder_path}/MatchupTable_updated.xlsx")