import sys
import json
from PyANGBasic import *
from PyANGKernel import *
from PyANGConsole import *


def main(argv):
    # Start a Console
    console = ANGConsole()
    # Load a network
    if console.open(argv[1]):
        model = console.getModel()

        import io
        import os
        import shutil
        import sys
        from pathlib import Path

        input_config = json.loads(argv[-1])
        _SITEPACKAGES = input_config["AIMSUN"]["site_packages"]

        # Point Aimsun's embedded Python at a site-packages that has pandas/numpy.
        # _SITEPACKAGES = r"C:\Users\ggx\AppData\Local\Programs\Python\Python37\Lib\site-packages"
        # _SITEPACKAGES = r"C:\Users\xh8\AppData\Local\Programs\Python\Python310\Lib\site-packages"

        if _SITEPACKAGES not in sys.path:
            sys.path.append(_SITEPACKAGES)

        import numpy as np
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Alignment

        def _default_output_dir():
            try:
                doc = str(model.getDocumentFileName())
            except Exception:
                doc = ""
            folder = os.path.dirname(doc) if doc else ""
            if not folder or not os.path.isdir(folder):
                folder = os.path.expanduser("~")
            return folder

        INPUT_DIR = _default_output_dir()

        MATCHUP_FILE = "MatchupTable.xlsx"   # raw table in INPUT_DIR, rewritten in place
        # sub-folder holding the Synchro UTDF file(s)
        CONTROL_SUBDIR = "Control"
        # sub-folder holding the GridSmart file(s)
        TRAFFIC_SUBDIR = "Traffic"

        # Save a .bak copy of the table before overwriting it.
        MAKE_BACKUP = True
        NETWORK_COLS = ["JunctionID_Aimsun", "Bearing", "Numbering",
                        "FromRoadID_Aimsun", "ToRoadID_Aimsun", "Turn"]
        DEMAND_COLS = ["File_GridSmart", "Date_GridSmart",
                       "IntersectionName_GridSmart", "Turn_GridSmart"]
        SIGNAL_COLS = ["File_Synchro",
                       "IntersectionID_Synchro", "Turn_Synchro"]
        OTHER_COLS = ["Need calibration?"]

        ALL_COLS = NETWORK_COLS + DEMAND_COLS + SIGNAL_COLS + OTHER_COLS

        # Network turn vocabulary -> GridSmart/Synchro movement letter.
        TURN_TO_LETTER = {"right": "R", "thru": "T", "left": "L", "Uturn": "U"}

        # Cardinal directions in compass (clockwise) order.
        CARDINAL_ORDER = ["NB", "NE", "EB", "SE", "SB", "SW", "WB", "NW"]

        def process_signal_from_utdf(file_utdf):
            """Process the signal data and return a dict of ``{table name: DataFrame}``."""
            SignalDict = {}
            current_table = None
            current_table_data = []

            with open(file_utdf, "r", encoding="utf-8") as f:
                file_lines = f.readlines()

            removal_flag = 0
            for line in file_lines:
                line = line.strip()

                # Check if it's a line to be skipped
                if removal_flag == 1:
                    removal_flag = 0
                    continue

                # Check if it's a table name (indicated by square brackets)
                if line.startswith("["):
                    removal_flag = 1
                    if current_table is None:
                        # Remove square brackets and ]
                        current_table = line[1:-1].split(",")[0].rstrip("]")
                    else:
                        # Store the previous table data in the dictionary
                        if current_table_data:
                            df = pd.read_csv(io.StringIO(
                                "\n".join(current_table_data)), dtype=str)
                            SignalDict[current_table] = df

                        # Start a new table
                        current_table = line[1:-1].split(",")[0].rstrip("]")
                        current_table_data = []
                else:
                    # Accumulate table data
                    current_table_data.append(line)

            # Store the last table in the dictionary
            if current_table_data:
                SignalDict[current_table] = pd.read_csv(
                    io.StringIO("\n".join(current_table_data)), dtype=str)

            return SignalDict

        def is_missing_or_zero(val):
            """Check if a value is missing or equivalent to zero (as a number or string)."""
            if pd.isna(val):
                return True
            val_str = str(val).strip()
            if val_str in {"", "0"}:
                return True
            try:
                if float(val_str) == 0:
                    return True
            except ValueError:
                pass
            return False

        def _align_turns_to_network(matchup_table, subset, column, present_directions,
                                    source_movements, source_label, warned_messages):
            """Fill ``column`` so it aligns with the network ``Turn`` column."""
            approaches = list(dict.fromkeys(subset["FromRoadID_Aimsun"]))
            approach_to_cardinal = dict(zip(approaches, present_directions))

            network_movements = set()
            for idx in subset.index:
                cardinal = approach_to_cardinal.get(
                    matchup_table.at[idx, "FromRoadID_Aimsun"])
                if cardinal is None:
                    continue
                code = cardinal + \
                    TURN_TO_LETTER.get(matchup_table.at[idx, "Turn"], "?")
                matchup_table.at[idx, column] = code
                network_movements.add(code)

            # Report turns present in one side but not the other (excluding U-turns).
            only_source = sorted(source_movements - network_movements)
            only_network = sorted(c for c in (network_movements - source_movements)
                                  if not c.endswith("U"))
            if only_source or only_network:
                detail = []
                if only_source:
                    detail.append(
                        f'in {source_label["side"]} but not network: {", ".join(only_source)}')
                if only_network:
                    detail.append(
                        f'in network but not {source_label["side"]}: {", ".join(only_network)}')
                msg = f'{source_label["message"]}\n  ({"; ".join(detail)})'
                if msg not in warned_messages:
                    warned_messages.add(msg)
                    print(msg)

        def generate_matchup_table(df_matchup_table, path_output="MatchUp_Table.xlsx"):
            """Write the matchup table to an Excel file with the 2-row merged header."""
            wb = Workbook()
            ws = wb.active

            ws.append(["Network"] * len(NETWORK_COLS) + ["Demand"] * len(DEMAND_COLS) +
                      ["Signal"] * len(SIGNAL_COLS) + [""] * len(OTHER_COLS))

            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1, end_column=len(NETWORK_COLS))
            ws.merge_cells(start_row=1, start_column=len(NETWORK_COLS) + 1, end_row=1,
                           end_column=len(NETWORK_COLS) + len(DEMAND_COLS))
            ws.merge_cells(start_row=1, start_column=len(NETWORK_COLS) + len(DEMAND_COLS) + 1,
                           end_row=1,
                           end_column=len(NETWORK_COLS) + len(DEMAND_COLS) + len(SIGNAL_COLS))

            ws.append(ALL_COLS)

            for row in df_matchup_table.itertuples(index=False):
                ws.append(list(row))

            current_start = 3  # Data starts at row 3
            for i in range(3, len(df_matchup_table) + 3):
                if (i == len(df_matchup_table) + 2 or ws["A%d" % i].value != ws["A%d" % (i + 1)].value):
                    if current_start < i:  # Only merge if there are multiple same values
                        # JunctionID
                        ws.merge_cells(start_row=current_start,
                                       start_column=1, end_row=i, end_column=1)
                        # File_GridSmart
                        ws.merge_cells(start_row=current_start,
                                       start_column=7, end_row=i, end_column=7)
                        # Date_GridSmart
                        ws.merge_cells(start_row=current_start,
                                       start_column=8, end_row=i, end_column=8)
                        # IntersectionName_GridSmart
                        ws.merge_cells(start_row=current_start,
                                       start_column=9, end_row=i, end_column=9)
                        # IntersectionID_Synchro
                        ws.merge_cells(start_row=current_start,
                                       start_column=12, end_row=i, end_column=12)
                        # Need calibration?
                        ws.merge_cells(start_row=current_start,
                                       start_column=14, end_row=i, end_column=14)
                    current_start = i + 1

            if len(df_matchup_table) > 0:
                ws.merge_cells(start_row=3, start_column=11,
                               end_row=len(df_matchup_table) + 2, end_column=11)

            # Center align merged cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center")

            column_widths = {"A": 20, "B": 15, "C": 15, "D": 25, "E": 25, "F": 15, "G": 20,
                             "H": 20, "I": 30, "J": 20, "K": 20, "L": 25, "M": 20, "N": 20}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            wb.save(path_output)
            return True

        def update_matchup_table(path_matchup_table, control_dir="", traffic_dir="",
                                 path_output=None):
            """Update the matchup table with GridSmart demand and Synchro UTDF data."""
            if path_output is None:
                path_output = path_matchup_table

            # read the lookup table
            MatchupTable_UserInput = pd.read_excel(
                path_matchup_table, skiprows=1, dtype=str)

            # Rename any legacy *_OpenDrive headers to *_Aimsun.
            legacy = {"JunctionID_OpenDrive": "JunctionID_Aimsun",
                      "FromRoadID_OpenDrive": "FromRoadID_Aimsun",
                      "ToRoadID_OpenDrive": "ToRoadID_Aimsun"}
            rename = {old: new for old, new in legacy.items()
                      if old in MatchupTable_UserInput.columns
                      and new not in MatchupTable_UserInput.columns}
            if rename:
                MatchupTable_UserInput = MatchupTable_UserInput.rename(columns=rename)
                print(f"  :Renamed legacy header(s) back to Aimsun naming: {', '.join(sorted(rename))}")

            missing = [
                c for c in ALL_COLS if c not in MatchupTable_UserInput.columns]
            if missing:
                raise ValueError(
                    "Matchup table {} is missing the column(s): {}\n"
                    "Expected the layout written by "
                    "aimsun_matchup_table_generation_aimsunid.py:\n  {}\n"
                    "Found:\n  {}".format(path_matchup_table, ", ".join(missing),
                       ", ".join(ALL_COLS), ", ".join(map(str, MatchupTable_UserInput.columns))))

            # Forward fill missing values in merged columns
            merged_columns = ["JunctionID_Aimsun", "File_Synchro"]
            MatchupTable_UserInput[merged_columns] = MatchupTable_UserInput[merged_columns].ffill(
            )
            MatchupTable_UserInput["Need calibration?"] = "Y"

            # Collect each distinct turn-mismatch warning so it is printed only once.
            warned_messages = set()

            # ---- Demand: fill Turn_GridSmart / IntersectionName / Date ---------------
            for junction_id in MatchupTable_UserInput["JunctionID_Aimsun"].unique():
                subset = MatchupTable_UserInput[
                    MatchupTable_UserInput["JunctionID_Aimsun"] == junction_id]

                file_name = (subset["File_GridSmart"].dropna().iloc[0]
                             if not subset["File_GridSmart"].isna().all() else None)
                if not file_name:
                    continue

                if "." not in Path(file_name).name:
                    file_name = file_name + ".xls"
                # A junction with counts does not need calibration.
                MatchupTable_UserInput.loc[
                    MatchupTable_UserInput["JunctionID_Aimsun"] == junction_id,
                    "Need calibration?"] = "N"

                gs_file_path = Path(traffic_dir) / file_name
                if not gs_file_path.suffix:  # suffix is empty if no extension
                    gs_file_path = gs_file_path.with_suffix(".xls")

                try:
                    gs_data = pd.read_excel(
                        gs_file_path, header=None, dtype=str)
                except FileNotFoundError:
                    print(f'GridSmart file not found, skipping: "{gs_file_path}"')
                    continue

                # Extract IntersectionName_GridSmart
                intersection_row = gs_data[gs_data.iloc[:, 0]
                                           == "Intersection"].index
                if not intersection_row.empty:
                    intersection_col = gs_data.iloc[intersection_row[0], 1:].first_valid_index(
                    )
                    if intersection_col is not None:
                        intersection_name = gs_data.iloc[intersection_row[0],
                                                         intersection_col]
                        MatchupTable_UserInput.loc[
                            MatchupTable_UserInput["JunctionID_Aimsun"] == junction_id,
                            "IntersectionName_GridSmart"] = intersection_name

                # Extract Date_GridSmart
                date_row = gs_data[gs_data.iloc[:, 0] == "Date"].index
                if not date_row.empty:
                    date_col = gs_data.iloc[date_row[0],
                                            1:].first_valid_index()
                    if date_col is not None:
                        date_value = gs_data.iloc[date_row[0], date_col]
                        MatchupTable_UserInput.loc[
                            MatchupTable_UserInput["JunctionID_Aimsun"] == junction_id,
                            "Date_GridSmart"] = date_value

                # Map each summary-block movement header to its column index.
                movement_label = {"Right": "R",
                                  "Through": "T", "Left": "L", "UTurn": "U"}
                movement_columns = {}
                for col in gs_data.columns[1:]:
                    for movement in movement_label:
                        if gs_data.iloc[:, col].eq(movement).any():
                            movement_columns[movement] = col

                # Collect, per cardinal direction, the turn letters that have data.
                demand_by_direction = {}
                for direction, prefix in zip(["Northbound", "Eastbound", "Southbound", "Westbound"],
                                             ["NB", "EB", "SB", "WB"]):
                    direction_row = gs_data[gs_data.iloc[:, 0]
                                            == direction].index
                    if direction_row.empty:
                        continue
                    movement_row = direction_row[0]
                    letters = [letter for movement, letter in movement_label.items()
                               if movement in movement_columns
                               and pd.notna(gs_data.iloc[movement_row, movement_columns[movement]])]
                    if letters:
                        demand_by_direction[prefix] = letters

                # Align Turn_GridSmart with the network Turn and report any difference.
                present_directions = [
                    d for d in CARDINAL_ORDER if d in demand_by_direction]
                demand_movements = {prefix + letter
                                    for prefix, letters in demand_by_direction.items()
                                    for letter in letters}
                _align_turns_to_network(
                    MatchupTable_UserInput, subset, "Turn_GridSmart",
                    present_directions, demand_movements,
                    {"side": "file",
                     "message": (f'Turn in demand file "{file_name}" is different from turn in '
                                 'network, please check.')},
                    warned_messages)

            # ---- Signal: fill Turn_Synchro ------------------------------------------
            synchro_cache = {}  # avoid re-reading the same UTDF file

            for junction_id in MatchupTable_UserInput["JunctionID_Aimsun"].unique():
                subset = MatchupTable_UserInput[
                    MatchupTable_UserInput["JunctionID_Aimsun"] == junction_id]

                file_synchro_name = (subset["File_Synchro"].dropna().iloc[0]
                                     if not subset["File_Synchro"].isna().all() else None)
                if subset["IntersectionID_Synchro"].dropna().empty:
                    continue

                intersection_id_synchro = subset["IntersectionID_Synchro"].dropna(
                ).iloc[0]
                if file_synchro_name in synchro_cache:
                    signal_dict = synchro_cache[file_synchro_name]
                else:
                    synchro_file_path = os.path.join(
                        control_dir, file_synchro_name)
                    try:
                        signal_dict = process_signal_from_utdf(
                            synchro_file_path)
                    except IOError:
                        print(f'Synchro file not found, skipping: "{synchro_file_path}"')
                        signal_dict = {}
                    synchro_cache[file_synchro_name] = signal_dict

                lanes_df = signal_dict.get("Lanes")
                if lanes_df is None:
                    print(f'No "Lanes" table found in "{file_synchro_name}".')
                    continue

                # Subset rows for this intersection with an allowed RECORDNAME.
                allowed_recordnames = ["Lanes", "Shared", "Phase1", "PermPhase1",
                                       "Phase2", "PermPhase2", "Phase3", "PermPhase3"]
                subset_lanes = lanes_df[
                    (lanes_df["INTID"].astype(str)
                     == str(intersection_id_synchro))
                    & (lanes_df["RECORDNAME"].astype(str).isin(allowed_recordnames))
                ].copy()

                for col in subset_lanes.columns:
                    if not col.endswith("T"):
                        continue

                    base = col[:-1]  # Extract base like 'XY' from 'XYT'
                    try:
                        phaseid = subset_lanes.loc[subset_lanes["RECORDNAME"]
                                                   == "Phase1", col].values[0]
                        shareid_val = subset_lanes.loc[subset_lanes["RECORDNAME"]
                                                       == "Shared", col].values[0]
                    except IndexError:
                        continue  # Skip if 'Phase1' or 'Shared' row not found
                    except KeyError:
                        continue  # Skip if col not in subset_lanes

                    if pd.isna(phaseid) or pd.isna(shareid_val):
                        continue

                    try:
                        shareid = int(float(shareid_val))
                    except ValueError:
                        continue

                    if shareid == 0:
                        continue

                    if shareid in [1, 3] and base + "L" in subset_lanes.columns:
                        # Left turn logic
                        l_col = base + "L"
                        l_sel = subset_lanes.loc[subset_lanes["RECORDNAME"]
                                                 == "Phase1", l_col]
                        l_perm_sel = subset_lanes.loc[subset_lanes["RECORDNAME"]
                                                      == "PermPhase1", l_col]
                        l_phase = l_sel.values[0] if not l_sel.empty else np.nan
                        l_perm = l_perm_sel.values[0] if not l_perm_sel.empty else np.nan
                        if pd.isna(l_phase) and pd.isna(l_perm):
                            subset_lanes.loc[subset_lanes["RECORDNAME"]
                                             == "PermPhase1", l_col] = phaseid

                    if shareid in [2, 3] and base + "R" in subset_lanes.columns:
                        # Right turn logic
                        r_col = base + "R"
                        r_sel = subset_lanes.loc[subset_lanes["RECORDNAME"]
                                                 == "Phase1", r_col]
                        r_perm_sel = subset_lanes.loc[subset_lanes["RECORDNAME"]
                                                      == "PermPhase1", r_col]
                        r_phase = r_sel.values[0] if not r_sel.empty else np.nan
                        r_perm = r_perm_sel.values[0] if not r_perm_sel.empty else np.nan
                        if pd.isna(r_phase) and pd.isna(r_perm):
                            subset_lanes.loc[subset_lanes["RECORDNAME"]
                                             == "Phase1", r_col] = phaseid

                if subset_lanes.empty:
                    print(f"No matching records in Lanes for IntersectionID_Synchro "
                          f"{intersection_id_synchro} in file {file_synchro_name}.")
                    continue

                subset_lanes.reset_index(drop=True, inplace=True)

                # Drop columns whose first-row value is missing/zero, unless an exception applies.
                cols_to_drop = []
                for col in subset_lanes.columns:
                    val_first = subset_lanes.at[0, col]
                    if not is_missing_or_zero(val_first):
                        continue

                    # Exception 1: a valid value from row 3 onward keeps the column.
                    if subset_lanes.shape[0] > 2:
                        subsequent_valid = subset_lanes[col].iloc[2:].apply(
                            lambda x: not is_missing_or_zero(x)).any()
                    else:
                        subsequent_valid = False

                    # Exception 2: for a column ending in 'R', check the matching XYT column.
                    exception2_keep = False
                    if col.endswith("R"):
                        col_t = col[:-1] + "T"
                        if col_t in subset_lanes.columns and subset_lanes.shape[0] > 1:
                            try:
                                val_first_t = float(subset_lanes.at[0, col_t])
                                val_second_t = float(subset_lanes.at[1, col_t])
                                if val_first_t > 0 and val_second_t > 1:
                                    exception2_keep = True
                            except ValueError:
                                pass

                    if not subsequent_valid and not exception2_keep:
                        cols_to_drop.append(col)

                subset_lanes.drop(columns=cols_to_drop, inplace=True)

                movements = [col for col in subset_lanes.columns
                             if col not in ["RECORDNAME", "INTID", "PED", "HOLD"]]

                # The Synchro Lanes columns are already cardinal movement codes (e.g. NBL).
                signal_movements = set(movements)
                present_directions = [d for d in CARDINAL_ORDER
                                      if any(m.startswith(d) for m in signal_movements)]

                _align_turns_to_network(
                    MatchupTable_UserInput, subset, "Turn_Synchro",
                    present_directions, signal_movements,
                    {"side": "signal",
                     "message": (f'Turn at intersection {intersection_id_synchro} '
                                 f'from signal file "{file_synchro_name}" is '
                                 'different from turn in network, please check.')},
                    warned_messages)

            # Write back with the column names exactly as they came in.
            generate_matchup_table(
                MatchupTable_UserInput[ALL_COLS], path_output)
            return MatchupTable_UserInput

        def _parse_command_line(argv):
            """Parse shell arguments."""
            import argparse

            parser = argparse.ArgumentParser(
                description="Fill the Demand/Signal columns of a raw Aimsun matchup table.")
            parser.add_argument("matchup_table",
                                help="path to the raw MatchupTable .xlsx")
            parser.add_argument("-c", "--control", default="Control",
                                help="directory holding the Synchro UTDF file(s) (default: Control)")
            parser.add_argument("-t", "--traffic", default="Traffic",
                                help="directory holding the GridSmart file(s) (default: Traffic)")
            parser.add_argument("-o", "--output", default=None,
                                help="write here instead of overwriting the input")
            parser.add_argument("--backup", action="store_true",
                                help="save a .bak copy of the input before overwriting it")
            args = parser.parse_args(argv)
            return (args.matchup_table, args.control, args.traffic, args.output, args.backup)

        def _launched_from_shell():
            """True only when python was pointed straight at this file."""
            try:
                this_file = os.path.basename(__file__)
            except NameError:
                return False        # pasted into the console; no file, no command line
            if not sys.argv:
                return False
            return os.path.basename(sys.argv[0]).lower() == this_file.lower()

        matchup = os.path.join(INPUT_DIR, MATCHUP_FILE)
        control = os.path.join(INPUT_DIR, CONTROL_SUBDIR)
        traffic = os.path.join(INPUT_DIR, TRAFFIC_SUBDIR)
        output = None
        backup = MAKE_BACKUP

        for label, path in (("matchup table", matchup),
                            ("control dir", control),
                            ("traffic dir", traffic)):
            if not os.path.exists(path):
                print(f"  :ERROR - {label} not found: {path}")
                print("  :Set INPUT_DIR at the top of the script and re-run.")
                return 1

        output = output or matchup
        if backup and os.path.abspath(output) == os.path.abspath(matchup):
            # Never overwrite an existing backup.
            backup_path = matchup + ".bak"
            n = 2
            while os.path.exists(backup_path):
                backup_path = f"{matchup}.bak{n}"
                n += 1
            shutil.copy2(matchup, backup_path)
            print(f"  :Backed up original to {backup_path}")

        print(f"  :Updating {matchup}")
        print(f"  :  control={control}")
        print(f"  :  traffic={traffic}")
        df = update_matchup_table(matchup, control, traffic, output)

        filled = df["Turn_GridSmart"].notna().sum()
        junctions = df.loc[df["File_GridSmart"].notna(),
                           "JunctionID_Aimsun"].nunique()
        print(f"  :Wrote {output}")
        print(f"  :{len(df)} row(s) total, {filled} with Turn_GridSmart, across {junctions} junction(s) "
              "with a GridSmart file.")
        if filled == 0:
            print("  :WARNING - no Turn_GridSmart was filled; check that File_GridSmart "
                  "is populated and that those files exist in the traffic directory.")
            return 1
        return 0

        console.save(argv[1])
        console.close()
    else:
        console.getLog().addError("Cannot load the network")
        print("cannot load network")


if __name__ == "__main__":
    sys.exit(main(sys.argv))
